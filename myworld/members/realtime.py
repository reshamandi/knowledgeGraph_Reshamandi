from time import sleep
from openpyxl import load_workbook
import os 

dir_path = os.path.dirname(os.path.realpath(__file__))
print (dir_path)

# while 1:
#     wb = load_workbook(filename = 'sample.numbers')
#     ws = wb['Sheet1']
#     print(ws.max_row)
#     sleep(1)